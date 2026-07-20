"""
relatorios.py
Monta o Excel de saida do Conciliador Contabil Inteligente, com 4 abas:
    Detalhe_Conciliacao, Resumo_Periodo, Itens_Em_Aberto, Ponte_Balancete

Convencoes seguidas (planilha financeira):
    - Fonte Arial em toda a pasta de trabalho
    - Somas sempre como formula do Excel (SUBTOTAL/SOMA), nunca valor fixo
    - Negativos entre parenteses; formatacao condicional por status
    - Cabecalho fixo (freeze panes) e colunas com largura ajustada
"""
from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger("conciliador.relatorios")

FONTE = "Arial"
FMT_MOEDA = '#,##0.00;[RED](#,##0.00)'
FMT_PCT = '0.0%'
FMT_DATA = 'dd/mm/yyyy'

COR_CABECALHO = "1F2937"
COR_TEXTO_CABECALHO = "FFFFFF"
COR_CONCILIADO = "E1F5EE"      # verde claro
COR_ABERTO_RECENTE = "FAEEDA"  # ambar claro (<=90 dias)
COR_ABERTO_ANTIGO = "FAECE7"   # coral claro (>90 dias)

_BORDA_FINA = Side(style="thin", color="D9D9D9")
BORDA_CELULA = Border(left=_BORDA_FINA, right=_BORDA_FINA, top=_BORDA_FINA, bottom=_BORDA_FINA)


def _estilo_cabecalho(ws: Worksheet, ultima_coluna: int, linha: int = 1) -> None:
    fill = PatternFill("solid", fgColor=COR_CABECALHO)
    fonte = Font(name=FONTE, bold=True, color=COR_TEXTO_CABECALHO, size=10)
    for col in range(1, ultima_coluna + 1):
        cel = ws.cell(row=linha, column=col)
        cel.fill = fill
        cel.font = fonte
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[linha].height = 32
    ws.freeze_panes = ws.cell(row=linha + 1, column=1)


def _escrever_dataframe(
    ws: Worksheet,
    df: pd.DataFrame,
    linha_inicial: int = 1,
    colunas_moeda: tuple[str, ...] = (),
    colunas_data: tuple[str, ...] = (),
    colunas_pct: tuple[str, ...] = (),
    larguras: dict | None = None,
) -> int:
    """Escreve um DataFrame a partir de linha_inicial, com cabecalho. Devolve a
    ultima linha escrita."""
    colunas = list(df.columns)
    for j, nome in enumerate(colunas, start=1):
        ws.cell(row=linha_inicial, column=j, value=nome.replace("_", " ").title())
    _estilo_cabecalho(ws, len(colunas), linha=linha_inicial)

    for i, (_, linha) in enumerate(df.iterrows(), start=linha_inicial + 1):
        for j, nome in enumerate(colunas, start=1):
            valor = linha[nome]
            if pd.isna(valor):
                valor = None
            cel = ws.cell(row=i, column=j, value=valor)
            cel.font = Font(name=FONTE, size=10)
            cel.border = BORDA_CELULA
            if nome in colunas_moeda:
                cel.number_format = FMT_MOEDA
            elif nome in colunas_data:
                cel.number_format = FMT_DATA
            elif nome in colunas_pct:
                cel.number_format = FMT_PCT

    larguras = larguras or {}
    for j, nome in enumerate(colunas, start=1):
        ws.column_dimensions[get_column_letter(j)].width = larguras.get(nome, 16)

    return linha_inicial + len(df)


def _cor_por_status(status: str, aging_dias: float | None) -> str | None:
    if str(status).startswith("Conciliado"):
        return COR_CONCILIADO
    if str(status).startswith("Em Aberto"):
        if aging_dias is not None and aging_dias > 90:
            return COR_ABERTO_ANTIGO
        return COR_ABERTO_RECENTE
    return None


def _aba_detalhe(wb: Workbook, resultado: pd.DataFrame) -> None:
    ws = wb.create_sheet("Detalhe_Conciliacao")
    colunas = [
        "id_lancamento", "linha_origem", "periodo", "data", "historico", "c_partida",
        "debito", "credito", "valor", "direcao", "tipo_textual",
        "status", "regra_aplicada", "id_match", "contraparte",
        "valor_residual", "aging_dias", "faixa_aging", "obs",
    ]
    df = resultado[colunas].sort_values(["periodo", "data"]).reset_index(drop=True)
    ultima_linha = _escrever_dataframe(
        ws, df,
        colunas_moeda=("debito", "credito", "valor", "valor_residual"),
        colunas_data=("data",),
        larguras={
            "id_lancamento": 12, "linha_origem": 10, "periodo": 9, "data": 12,
            "historico": 42, "c_partida": 14, "debito": 13, "credito": 13,
            "valor": 13, "direcao": 16, "tipo_textual": 14, "status": 22,
            "regra_aplicada": 16, "id_match": 10, "contraparte": 26,
            "valor_residual": 14, "aging_dias": 11, "faixa_aging": 13, "obs": 14,
        },
    )
    idx_status = colunas.index("status") + 1
    idx_aging = colunas.index("aging_dias") + 1
    for i, row in enumerate(df.itertuples(index=False), start=2):
        cor = _cor_por_status(getattr(row, "status"), getattr(row, "aging_dias"))
        if cor:
            for col in range(1, len(colunas) + 1):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=cor)

    # linha de totais com SUBTOTAL (ignora linhas ocultas/filtradas se o usuario filtrar)
    linha_total = ultima_linha + 1
    ws.cell(row=linha_total, column=1, value="Total").font = Font(name=FONTE, bold=True)
    for nome_col in ("debito", "credito", "valor", "valor_residual"):
        j = colunas.index(nome_col) + 1
        letra = get_column_letter(j)
        cel = ws.cell(row=linha_total, column=j, value=f"=SUBTOTAL(9,{letra}2:{letra}{ultima_linha})")
        cel.number_format = FMT_MOEDA
        cel.font = Font(name=FONTE, bold=True)

    linha_legenda = linha_total + 1
    j_valor = colunas.index("valor") + 1
    j_residual = colunas.index("valor_residual") + 1
    fonte_legenda = Font(name=FONTE, italic=True, size=9, color="6B7280")
    ws.cell(
        row=linha_legenda, column=j_valor,
        value="Valor: lançamento original (não muda com a conciliação)",
    ).font = fonte_legenda
    ws.cell(
        row=linha_legenda, column=j_residual,
        value="Valor_Residual: saldo real em aberto (zera quando Obs = \"efeito zero\")",
    ).font = fonte_legenda

    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}{ultima_linha}"


def _aba_resumo(wb: Workbook, resumo: pd.DataFrame) -> None:
    ws = wb.create_sheet("Resumo_Periodo")
    ultima_linha = _escrever_dataframe(
        ws, resumo,
        colunas_moeda=("total_provisionado", "total_revertido_baixado", "saldo_liquido_periodo", "saldo_em_aberto_no_periodo"),
        larguras={
            "periodo": 10, "total_provisionado": 18, "total_revertido_baixado": 20,
            "saldo_liquido_periodo": 18, "saldo_em_aberto_no_periodo": 20, "pct_conciliado": 14,
        },
    )
    # % conciliado como fracao para o formato de percentual funcionar (0-100 -> 0.0-1.0)
    col_pct = list(resumo.columns).index("pct_conciliado") + 1
    for i in range(2, ultima_linha + 1):
        cel = ws.cell(row=i, column=col_pct)
        if cel.value is not None:
            cel.value = cel.value / 100
            cel.number_format = FMT_PCT

    linha_total = ultima_linha + 1
    ws.cell(row=linha_total, column=1, value="Total").font = Font(name=FONTE, bold=True)
    for nome_col in ("total_provisionado", "total_revertido_baixado", "saldo_liquido_periodo", "saldo_em_aberto_no_periodo"):
        j = list(resumo.columns).index(nome_col) + 1
        letra = get_column_letter(j)
        cel = ws.cell(row=linha_total, column=j, value=f"=SUM({letra}2:{letra}{ultima_linha})")
        cel.number_format = FMT_MOEDA
        cel.font = Font(name=FONTE, bold=True)


def _aba_itens_abertos(wb: Workbook, resultado: pd.DataFrame) -> None:
    ws = wb.create_sheet("Itens_Em_Aberto")
    abertos = resultado[resultado["residual_centavos"] != 0].copy()
    abertos = abertos.sort_values("aging_dias", ascending=False)
    colunas = [
        "id_lancamento", "periodo", "data", "historico", "c_partida",
        "valor", "valor_residual", "status", "aging_dias", "faixa_aging",
    ]
    if abertos.empty:
        ws.cell(row=1, column=1, value="Nenhum item em aberto - conta 100% conciliada.")
        return
    ultima_linha = _escrever_dataframe(
        ws, abertos[colunas],
        colunas_moeda=("valor", "valor_residual"),
        colunas_data=("data",),
        larguras={
            "id_lancamento": 12, "periodo": 9, "data": 12, "historico": 42,
            "c_partida": 14, "valor": 13, "valor_residual": 14, "status": 20,
            "aging_dias": 11, "faixa_aging": 13,
        },
    )
    linha_total = ultima_linha + 1
    ws.cell(row=linha_total, column=1, value="Total em aberto").font = Font(name=FONTE, bold=True)
    j = colunas.index("valor_residual") + 1
    letra = get_column_letter(j)
    cel = ws.cell(row=linha_total, column=j, value=f"=SUM({letra}2:{letra}{ultima_linha})")
    cel.number_format = FMT_MOEDA
    cel.font = Font(name=FONTE, bold=True)


def _aba_ponte(wb: Workbook, ponte: pd.DataFrame, conta: str, periodo_referencia: str) -> None:
    ws = wb.create_sheet("Ponte_Balancete")
    ws.cell(row=1, column=1, value=f"Ponte de diferenças — conta {conta} — {periodo_referencia}").font = Font(
        name=FONTE, bold=True, size=12
    )
    ws.merge_cells("A1:C1")
    _escrever_dataframe(ws, ponte, linha_inicial=3, colunas_moeda=("valor",),
                         larguras={"item": 55, "valor": 16})
    ws.cell(row=2, column=1, value=(
        "(1) vem do Balancete oficial · (2) é a soma dos Itens em Aberto após a "
        "conciliação automática · (3) é o que ainda precisa de explicação manual "
        "antes do fechamento do mês."
    )).font = Font(name=FONTE, italic=True, size=9, color="6B7280")


def gerar_excel_saida(
    resultado: pd.DataFrame,
    resumo: pd.DataFrame,
    ponte: pd.DataFrame,
    caminho_saida: str | Path,
    conta: str = "",
    periodo_referencia: str = "",
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)  # remove a aba padrao vazia

    _aba_detalhe(wb, resultado)
    _aba_resumo(wb, resumo)
    _aba_itens_abertos(wb, resultado)
    _aba_ponte(wb, ponte, conta, periodo_referencia)

    caminho_saida = Path(caminho_saida)
    wb.save(caminho_saida)
    logger.info("Excel de saída gravado em %s", caminho_saida)
    return caminho_saida
