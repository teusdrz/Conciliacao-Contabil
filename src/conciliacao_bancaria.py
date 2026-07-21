"""
conciliacao_bancaria.py
Conciliacao bancaria linha a linha: "01.Razao" (razao contabil) x "02.Financeiro"
(extrato bancario / documento suporte).

Diferenca em relacao ao motor_conciliacao.py: aquele modulo casa o razao CONTRA
ELE MESMO (procura pares/grupos de lancamentos que se cancelam). Este modulo
casa o razao contra uma FONTE EXTERNA (o extrato bancario da aba "02.Financeiro"),
lancamento a lancamento, pela conta bancaria + valor + data (com tolerancia).

Nao importa nem modifica nada de motor_conciliacao.py/excel_io.py/main.py - e um
metodo novo e independente, so reaproveita os estilos visuais de relatorios.py
para manter a planilha de saida com a mesma cara do resto do projeto.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from relatorios import (
    BORDA_CELULA,
    COR_ABERTO_ANTIGO,
    COR_CONCILIADO,
    FMT_DATA,
    FMT_MOEDA,
    FONTE,
    _estilo_cabecalho,
)

logger = logging.getLogger("conciliador.conciliacao_bancaria")

ABA_RAZAO_PADRAO = "01.Razão"
ABA_FINANCEIRO_PADRAO = "02.Financeiro"
# Nomes de aba do Excel tem limite de 31 caracteres - "04.Conciliação Razão x
# Financeiro" (34) estourava esse limite e o Excel/Google Sheets recusa ou
# corrompe o nome da aba nesse caso.
ABA_SAIDA_PADRAO = "04.Razão x Financeiro"

STATUS_CONCILIADO = "Conciliado"
STATUS_SO_RAZAO = "Só no Razão"
STATUS_SO_FINANCEIRO = "Só no Financeiro"


def _parece_conta_bancaria(valor) -> bool:
    """Filtro simples pra identificar linhas de dado real (nao cabecalho/rodape
    de bloco) na aba 02.Financeiro: o codigo de conta contabil sempre vem no
    formato 'X.X.X.XX.XXXX'."""
    return isinstance(valor, str) and "." in valor and valor[:1].isdigit()


def _parse_data(valor) -> datetime | None:
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(valor.strip(), fmt)
            except ValueError:
                continue
    return None


@dataclass
class ResultadoConciliacaoBancaria:
    df: pd.DataFrame
    contas: list[str]
    periodo_inicio: datetime
    periodo_fim: datetime
    total_razao: float = field(init=False)
    total_financeiro: float = field(init=False)
    total_diferenca: float = field(init=False)
    qtd_conciliado: int = field(init=False)
    qtd_so_razao: int = field(init=False)
    qtd_so_financeiro: int = field(init=False)

    def __post_init__(self) -> None:
        self.total_razao = round(float(self.df["valor_razao"].fillna(0).sum()), 2)
        self.total_financeiro = round(float(self.df["valor_financeiro"].fillna(0).sum()), 2)
        self.total_diferenca = round(float(self.df["diferenca"].fillna(0).sum()), 2)
        self.qtd_conciliado = int((self.df["status"] == STATUS_CONCILIADO).sum())
        self.qtd_so_razao = int((self.df["status"] == STATUS_SO_RAZAO).sum())
        self.qtd_so_financeiro = int((self.df["status"] == STATUS_SO_FINANCEIRO).sum())


def ler_financeiro(arquivo: str | Path, aba: str = ABA_FINANCEIRO_PADRAO) -> pd.DataFrame:
    """Le a aba 02.Financeiro (extrato bancario, organizado em blocos por conta,
    cada bloco com cabecalho/rodape proprio). Em vez de tentar mapear os limites
    de cada bloco, identifica linha de dado real por ter conta + data validas -
    robusto ao layout de blocos e nao depende da posicao exata de cada um."""
    wb = load_workbook(str(arquivo), data_only=True)
    ws = wb[aba]

    linhas = []
    for row in range(1, ws.max_row + 1):
        conta = ws.cell(row=row, column=3).value
        data_raw = ws.cell(row=row, column=4).value
        if not _parece_conta_bancaria(conta):
            continue
        data = _parse_data(data_raw)
        if data is None:
            continue

        banco = ws.cell(row=row, column=2).value
        operacao = ws.cell(row=row, column=5).value
        documento = ws.cell(row=row, column=6).value
        entradas = ws.cell(row=row, column=8).value or 0
        saidas = ws.cell(row=row, column=9).value or 0
        movimentacao = ws.cell(row=row, column=11).value
        if not isinstance(movimentacao, (int, float)):
            movimentacao = (entradas or 0) - (saidas or 0)
        obs = ws.cell(row=row, column=12).value

        linhas.append(
            {
                "conta": str(conta).strip(),
                "banco": banco,
                "data": data,
                "operacao": operacao,
                "documento": documento,
                "valor": round(float(movimentacao), 2),
                "obs": obs,
                "linha_origem": row,
            }
        )

    df = pd.DataFrame(linhas)
    logger.info("Lidas %d linha(s) de transacao em '%s' (aba %s).", len(df), arquivo, aba)
    return df


def ler_razao_contas_bancarias(
    arquivo: str | Path,
    periodos_por_conta: dict[str, tuple[datetime, datetime]],
    aba: str = ABA_RAZAO_PADRAO,
    linha_cabecalho: int = 5,
) -> pd.DataFrame:
    """Le a aba 01.Razao e filtra so as linhas cujas contas aparecem no extrato
    (02.Financeiro) e cuja data cai dentro do periodo coberto pelo extrato
    DAQUELA MESMA CONTA especificamente (nao um periodo global entre todas as
    contas - contas diferentes podem ter janelas de datas diferentes no
    extrato, e usar um periodo global inflaria "só no Razão" com lançamentos
    de meses que aquela conta nem tem dado de extrato pra comparar)."""
    wb = load_workbook(str(arquivo), data_only=True)
    ws = wb[aba]

    linhas = []
    for row in range(linha_cabecalho + 1, ws.max_row + 1):
        conta = ws.cell(row=row, column=3).value
        conta = str(conta).strip() if conta is not None else None
        if conta not in periodos_por_conta:
            continue
        data_min, data_max = periodos_por_conta[conta]
        data = ws.cell(row=row, column=5).value
        if not isinstance(data, datetime):
            data = _parse_data(data)
        if data is None or not (data_min <= data <= data_max):
            continue

        valor = ws.cell(row=row, column=11).value or 0
        linhas.append(
            {
                "conta": str(conta).strip(),
                "periodo": ws.cell(row=row, column=2).value,
                "data": data,
                "documento": ws.cell(row=row, column=6).value,
                "historico": ws.cell(row=row, column=7).value,
                "valor": round(float(valor), 2),
                "linha_origem": row,
            }
        )

    df = pd.DataFrame(linhas)
    logger.info(
        "Lidas %d linha(s) do razão em '%s' (aba %s) para %d conta(s) bancária(s) (período por conta).",
        len(df), arquivo, aba, len(periodos_por_conta),
    )
    return df


def conciliar_razao_financeiro(
    df_razao: pd.DataFrame,
    df_financeiro: pd.DataFrame,
    tolerancia_valor: float = 0.01,
    tolerancia_dias: int = 5,
) -> pd.DataFrame:
    """Casa 1:1 os lancamentos do razao com os do extrato financeiro, por conta
    bancaria + valor (com tolerancia de centavos) + data mais proxima (com
    tolerancia de dias, pois o razao e o extrato podem registrar o mesmo
    lancamento em dias diferentes - ex.: compensacao vs. lancamento contabil).
    O que sobrar sem par vira diferenca (Só no Razão / Só no Financeiro)."""
    tol_cent = round(tolerancia_valor * 100)
    linhas_resultado: list[dict] = []

    contas = sorted(set(df_razao["conta"]).union(df_financeiro["conta"]))
    for conta in contas:
        razao_conta = df_razao[df_razao["conta"] == conta].reset_index(drop=True)
        fin_conta = df_financeiro[df_financeiro["conta"] == conta].reset_index(drop=True)

        candidatos = []
        for i, r in razao_conta.iterrows():
            valor_r_cent = round(r["valor"] * 100)
            for j, f in fin_conta.iterrows():
                valor_f_cent = round(f["valor"] * 100)
                if abs(valor_r_cent - valor_f_cent) > tol_cent:
                    continue
                diff_dias = abs((r["data"] - f["data"]).days)
                if diff_dias > tolerancia_dias:
                    continue
                candidatos.append((diff_dias, abs(valor_r_cent - valor_f_cent), i, j))

        candidatos.sort(key=lambda c: (c[0], c[1]))
        usados_razao: set[int] = set()
        usados_fin: set[int] = set()
        pares: list[tuple[int, int]] = []
        for _, _, i, j in candidatos:
            if i in usados_razao or j in usados_fin:
                continue
            usados_razao.add(i)
            usados_fin.add(j)
            pares.append((i, j))

        for i, j in pares:
            r = razao_conta.loc[i]
            f = fin_conta.loc[j]
            linhas_resultado.append(
                {
                    "conta": conta,
                    "data_razao": r["data"],
                    "historico_razao": r["historico"],
                    "documento_razao": r["documento"],
                    "valor_razao": r["valor"],
                    "data_financeiro": f["data"],
                    "operacao_financeiro": f["operacao"],
                    "documento_financeiro": f["documento"],
                    "valor_financeiro": f["valor"],
                    "diferenca": round(r["valor"] - f["valor"], 2),
                    "status": STATUS_CONCILIADO,
                    "linha_razao": r["linha_origem"],
                    "linha_financeiro": f["linha_origem"],
                }
            )

        for i, r in razao_conta.iterrows():
            if i in usados_razao:
                continue
            linhas_resultado.append(
                {
                    "conta": conta,
                    "data_razao": r["data"],
                    "historico_razao": r["historico"],
                    "documento_razao": r["documento"],
                    "valor_razao": r["valor"],
                    "data_financeiro": None,
                    "operacao_financeiro": None,
                    "documento_financeiro": None,
                    "valor_financeiro": 0.0,
                    "diferenca": r["valor"],
                    "status": STATUS_SO_RAZAO,
                    "linha_razao": r["linha_origem"],
                    "linha_financeiro": None,
                }
            )

        for j, f in fin_conta.iterrows():
            if j in usados_fin:
                continue
            linhas_resultado.append(
                {
                    "conta": conta,
                    "data_razao": None,
                    "historico_razao": None,
                    "documento_razao": None,
                    "valor_razao": 0.0,
                    "data_financeiro": f["data"],
                    "operacao_financeiro": f["operacao"],
                    "documento_financeiro": f["documento"],
                    "valor_financeiro": f["valor"],
                    "diferenca": -f["valor"],
                    "status": STATUS_SO_FINANCEIRO,
                    "linha_razao": None,
                    "linha_financeiro": f["linha_origem"],
                }
            )

    df = pd.DataFrame(linhas_resultado)
    if not df.empty:
        df = df.sort_values(["conta", "data_razao", "data_financeiro"], na_position="last").reset_index(drop=True)
    logger.info(
        "Conciliação concluída: %d linha(s) (%d conciliadas, %d só no razão, %d só no financeiro).",
        len(df),
        int((df["status"] == STATUS_CONCILIADO).sum()) if not df.empty else 0,
        int((df["status"] == STATUS_SO_RAZAO).sum()) if not df.empty else 0,
        int((df["status"] == STATUS_SO_FINANCEIRO).sum()) if not df.empty else 0,
    )
    return df


def executar(
    arquivo: str | Path,
    arquivo_saida: str | Path,
    aba_razao: str = ABA_RAZAO_PADRAO,
    aba_financeiro: str = ABA_FINANCEIRO_PADRAO,
    linha_cabecalho_razao: int = 5,
    nome_aba_saida: str = ABA_SAIDA_PADRAO,
    tolerancia_valor: float = 0.01,
    tolerancia_dias: int = 5,
) -> ResultadoConciliacaoBancaria:
    """Orquestra a leitura das duas abas, a conciliação e a gravação da aba nova
    de resultado numa cópia do arquivo original (nenhuma aba/fórmula/dado
    existente é alterado)."""
    df_financeiro = ler_financeiro(arquivo, aba=aba_financeiro)
    contas = set(df_financeiro["conta"])
    data_min = df_financeiro["data"].min()
    data_max = df_financeiro["data"].max()

    # periodo de cada conta e' o proprio min/max de datas do extrato DAQUELA
    # conta - contas diferentes podem ter janelas diferentes.
    periodos_por_conta = {
        conta: (grupo["data"].min(), grupo["data"].max())
        for conta, grupo in df_financeiro.groupby("conta")
    }

    df_razao = ler_razao_contas_bancarias(
        arquivo, periodos_por_conta=periodos_por_conta,
        aba=aba_razao, linha_cabecalho=linha_cabecalho_razao,
    )

    df_resultado = conciliar_razao_financeiro(
        df_razao, df_financeiro, tolerancia_valor=tolerancia_valor, tolerancia_dias=tolerancia_dias,
    )

    resultado = ResultadoConciliacaoBancaria(
        df=df_resultado, contas=sorted(contas), periodo_inicio=data_min, periodo_fim=data_max,
    )

    _gravar_aba_conciliacao(arquivo, resultado, arquivo_saida, nome_aba=nome_aba_saida)
    return resultado


def _gravar_aba_conciliacao(
    arquivo_entrada: str | Path,
    resultado: ResultadoConciliacaoBancaria,
    arquivo_saida: str | Path,
    nome_aba: str = ABA_SAIDA_PADRAO,
) -> Path:
    """Devolve o PROPRIO arquivo original (todas as abas, fórmulas e formatação
    intactas) só ACRESCENTANDO a aba nova de conciliação. Se essa aba (que é
    sempre gerada por este mesmo script) já existir de uma execução anterior,
    ela é substituída - nenhuma outra aba é tocada."""
    if len(nome_aba) > 31:
        nome_aba_original, nome_aba = nome_aba, nome_aba[:31]
        logger.warning(
            "Nome de aba '%s' tem mais de 31 caracteres (limite do Excel); truncado para '%s'.",
            nome_aba_original, nome_aba,
        )
    wb = load_workbook(str(arquivo_entrada))
    if nome_aba in wb.sheetnames:
        del wb[nome_aba]
    ws = wb.create_sheet(nome_aba)

    linha = 1
    ws.cell(row=linha, column=1, value="Conciliação Bancária - 01.Razão x 02.Financeiro").font = Font(
        name=FONTE, bold=True, size=13,
    )
    linha += 1
    periodo_txt = (
        f"Período consolidado: {resultado.periodo_inicio:%d/%m/%Y} a {resultado.periodo_fim:%d/%m/%Y} "
        "(cada conta é comparada só dentro da própria janela de datas do extrato)"
    )
    ws.cell(row=linha, column=1, value=periodo_txt).font = Font(name=FONTE, italic=True, size=10)
    linha += 2

    resumo = [
        ("Total Razão", resultado.total_razao, True),
        ("Total Financeiro", resultado.total_financeiro, True),
        ("Diferença total", resultado.total_diferenca, True),
        ("Qtd. conciliado", resultado.qtd_conciliado, False),
        ("Qtd. só no Razão", resultado.qtd_so_razao, False),
        ("Qtd. só no Financeiro", resultado.qtd_so_financeiro, False),
    ]
    for rotulo, valor, eh_moeda in resumo:
        ws.cell(row=linha, column=1, value=rotulo).font = Font(name=FONTE, bold=True, size=10)
        cel_valor = ws.cell(row=linha, column=2, value=valor)
        cel_valor.font = Font(name=FONTE, size=10)
        if eh_moeda:
            cel_valor.number_format = FMT_MOEDA
        linha += 1
    linha += 1

    colunas = [
        ("conta", "Conta", 16, None),
        ("data_razao", "Data Razão", 13, FMT_DATA),
        ("historico_razao", "Histórico Razão", 38, None),
        ("documento_razao", "Documento Razão", 20, None),
        ("valor_razao", "Valor Razão", 14, FMT_MOEDA),
        ("data_financeiro", "Data Financeiro", 15, FMT_DATA),
        ("operacao_financeiro", "Operação Financeiro", 26, None),
        ("documento_financeiro", "Documento Financeiro", 18, None),
        ("valor_financeiro", "Valor Financeiro", 15, FMT_MOEDA),
        ("diferenca", "Diferença", 13, FMT_MOEDA),
        ("status", "Status", 18, None),
    ]
    linha_cabecalho = linha
    for j, (_, titulo, _largura, _fmt) in enumerate(colunas, start=1):
        ws.cell(row=linha_cabecalho, column=j, value=titulo)
    _estilo_cabecalho(ws, len(colunas), linha=linha_cabecalho)

    fill_conciliado = PatternFill("solid", fgColor=COR_CONCILIADO)
    fill_diferenca = PatternFill("solid", fgColor=COR_ABERTO_ANTIGO)

    for i, (_, linha_dados) in enumerate(resultado.df.iterrows(), start=linha_cabecalho + 1):
        status = linha_dados["status"]
        for j, (chave, _titulo, _largura, fmt) in enumerate(colunas, start=1):
            valor = linha_dados[chave]
            if pd.isna(valor):
                valor = None
            cel = ws.cell(row=i, column=j, value=valor)
            cel.font = Font(name=FONTE, size=10)
            cel.border = BORDA_CELULA
            if fmt:
                cel.number_format = fmt
            cel.fill = fill_conciliado if status == STATUS_CONCILIADO else fill_diferenca

    for j, (_, _titulo, largura, _fmt) in enumerate(colunas, start=1):
        ws.column_dimensions[get_column_letter(j)].width = largura

    ws.freeze_panes = ws.cell(row=linha_cabecalho + 1, column=1)

    caminho_saida = Path(arquivo_saida)
    wb.save(caminho_saida)
    logger.info(
        "Aba '%s' gravada em %s (%d linha(s), todas as demais abas preservadas).",
        nome_aba, caminho_saida, len(resultado.df),
    )
    return caminho_saida
