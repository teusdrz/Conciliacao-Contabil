"""
excel_io.py
Leitura do razao contabil (aba estilo "01.Razao (2)") e escrita do Excel de saida.

Convencao de sinal (a mesma do arquivo original do usuario):
    Valor = DEBITO - CREDITO
    Valor > 0  -> lancamento a DEBITO  -> "saida" (baixa/reversao, reduz o passivo)
    Valor < 0  -> lancamento a CREDITO -> "entrada" (provisao, aumenta o passivo)
"""
from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

logger = logging.getLogger("conciliador.excel_io")

# Nomes de coluna esperados na linha de cabecalho da planilha de origem
# (na ordem em que aparecem, a partir da coluna B - a coluna A e sempre vazia
# nos arquivos exportados do ERP que motivaram este projeto)
COLUNAS_ORIGEM = [
    "periodo",
    "conta",
    "conta_desc",
    "data",
    "lote",
    "historico",
    "c_partida",
    "fl",
    "c_custo",
    "debito",
    "credito",
    "valor",
    "obs",
]


@dataclass
class RazaoCarregado:
    df: pd.DataFrame
    total_valor_calculado: float
    total_valor_cache_formula: float | None
    linha_cabecalho: int
    aba: str
    arquivo: str

    @property
    def diferenca_cache(self) -> float:
        """Diferenca entre o total recalculado em Python e o valor em cache da
        formula SUBTOTAL do Excel. Se != 0, o cache da planilha esta desatualizado
        (isso aconteceu no arquivo de referencia deste projeto - ver README)."""
        if self.total_valor_cache_formula is None:
            return 0.0
        return round(self.total_valor_calculado - self.total_valor_cache_formula, 2)


def carregar_razao(
    arquivo: str | Path,
    aba: str = "01.Razão (2)",
    linha_cabecalho: int = 6,
) -> RazaoCarregado:
    """Le a aba de razao contabil e devolve um DataFrame limpo + o total de
    conferencia.

    linha_cabecalho e 1-indexado (como o usuario ve no Excel). Os dados comecam
    na linha seguinte. Linhas sem data sao descartadas (cobre linhas de titulo,
    linhas de total e linhas em branco no fim da planilha).
    """
    arquivo = str(arquivo)
    header_idx = linha_cabecalho - 1  # pandas usa indice 0

    bruto = pd.read_excel(arquivo, sheet_name=aba, header=header_idx)

    # A primeira coluna (A) e sempre vazia nos arquivos de origem deste projeto;
    # as 13 colunas uteis comecam em B. Se o layout do usuario for diferente,
    # ajuste aqui - e o UNICO lugar que precisa saber da geometria da planilha.
    bruto = bruto.iloc[:, 1: 1 + len(COLUNAS_ORIGEM)].copy()
    bruto.columns = COLUNAS_ORIGEM

    # guarda o numero da linha original do Excel para rastreabilidade/auditoria
    bruto["linha_origem"] = bruto.index + linha_cabecalho + 1

    antes = len(bruto)
    df = bruto.dropna(subset=["data"]).copy()
    logger.info("Lidas %d linhas brutas, %d com data valida (dados reais).", antes, len(df))

    # tipos
    df["data"] = pd.to_datetime(df["data"])
    for col in ("debito", "credito", "valor"):
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    for col in ("periodo", "conta", "conta_desc", "lote", "historico", "c_partida", "fl", "c_custo", "obs"):
        df[col] = df[col].astype("string").fillna("")

    # normaliza periodo para int quando possivel (vem como float/str do Excel)
    def _periodo_int(v: str):
        try:
            return int(float(v))
        except ValueError:
            return None

    df["periodo"] = df["periodo"].map(_periodo_int)

    # Valor deve ser sempre debito - credito. Se a planilha de origem trouxer
    # um Valor divergente disso (erro de digitacao/formula quebrada), recalcula
    # e loga um aviso em vez de confiar cegamente na coluna.
    valor_esperado = df["debito"] - df["credito"]
    divergentes = (df["valor"] - valor_esperado).abs() > 0.005
    if divergentes.any():
        logger.warning(
            "%d linha(s) com Valor != Debito-Credito na planilha original (linhas Excel: %s). "
            "Recalculando Valor a partir de Debito/Credito.",
            divergentes.sum(),
            df.loc[divergentes, "linha_origem"].tolist(),
        )
    df["valor"] = valor_esperado

    df = df.reset_index(drop=True)
    df["id_lancamento"] = df.index.map(lambda i: f"L{i + 1:04d}")

    total_calculado = round(float(df["valor"].sum()), 2)
    total_cache = _ler_valor_cache_subtotal(arquivo, aba)

    resultado = RazaoCarregado(
        df=df,
        total_valor_calculado=total_calculado,
        total_valor_cache_formula=total_cache,
        linha_cabecalho=linha_cabecalho,
        aba=aba,
        arquivo=arquivo,
    )

    if resultado.diferenca_cache != 0:
        logger.warning(
            "O total recalculado (%.2f) difere do valor em cache da formula "
            "SUBTOTAL da planilha original (%.2f). O cache do Excel provavelmente "
            "esta desatualizado (a formula nao foi recalculada apos a ultima edicao) "
            "- o sistema usa sempre o valor recalculado em Python, nunca o cache.",
            total_calculado,
            total_cache,
        )
    return resultado


def atualizar_obs_arquivo_original(
    arquivo_original: str | Path,
    resultado: pd.DataFrame,
    caminho_saida: str | Path,
    aba: str = "01.Razão (2)",
    coluna_obs: str = "N",
) -> Path:
    """Devolve o PROPRIO arquivo original (todas as abas, formulas, formatacao e
    dados intactos) mudando apenas o conteudo da coluna Obs. da aba informada.

    Nenhuma outra aba e nenhuma outra coluna dessa aba e tocada - inclusive a
    formula de conferencia (=SUBTOTAL(...)) perto do topo continua exatamente
    como estava. So se escreve "Efeito zero" quando o grupo daquela linha
    realmente fecha em zero (mesma regra do motor de conciliacao); caso
    contrario a celula fica em branco, igual a convencao ja usada no arquivo
    original para itens que nao fecham sozinhos.
    """
    wb = load_workbook(arquivo_original)
    ws = wb[aba]
    col_idx = column_index_from_string(coluna_obs)

    for linha_origem, obs in zip(resultado["linha_origem"], resultado["obs"]):
        texto = "Efeito zero" if str(obs).strip().lower() == "efeito zero" else None
        # ws.cell(..., value=None) NAO limpa a celula (openpyxl trata None como
        # "valor nao informado", nao como "apagar") - por isso a atribuicao
        # direta de .value e obrigatoria aqui.
        ws.cell(row=int(linha_origem), column=col_idx).value = texto

    caminho_saida = Path(caminho_saida)
    wb.save(caminho_saida)
    logger.info(
        "Arquivo original atualizado (só a coluna %s da aba '%s') salvo em %s",
        coluna_obs, aba, caminho_saida,
    )
    return caminho_saida


def ler_saldo_balancete(
    arquivo: str | Path,
    conta: str,
    aba: str = "00.Balancete",
    linha_cabecalho: int = 7,
    coluna_conta: str = "Conta",
    coluna_saldo: str = "Obs.",
) -> float | None:
    """Busca o saldo assinado (credito negativo, debito positivo) de uma conta em
    uma aba de balancete no layout comum de export do Totvs Protheus: cabecalho
    com colunas repetidas (ex.: '31.12.25' aparece uma vez sem sinal e de novo,
    mais adiante na mesma linha, ja assinada). A funcao pega a ULTIMA ocorrencia
    de uma coluna numerica logo antes de 'Obs.' - que e a coluna assinada do
    periodo mais recente no layout observado.

    Isso e um atalho especifico para esse formato de balancete - se o seu vier
    diferente, prefira passar o saldo manualmente para ponte_balancete().
    Devolve None se nao conseguir localizar a conta ou o layout esperado.
    """
    try:
        wb = load_workbook(arquivo, read_only=True, data_only=True)
        ws = wb[aba]
        linhas = list(ws.iter_rows(values_only=True))
        cabecalho = linhas[linha_cabecalho - 1]
        idx_conta = cabecalho.index(coluna_conta)
        # a ultima coluna numerica antes de 'Obs.' e a que queremos
        idx_obs = cabecalho.index(coluna_saldo) if coluna_saldo in cabecalho else len(cabecalho)
        idx_saldo = idx_obs - 1
        for linha in linhas[linha_cabecalho:]:
            if linha[idx_conta] == conta:
                valor = linha[idx_saldo]
                return float(valor) if isinstance(valor, (int, float)) else None
    except Exception as exc:  # noqa: BLE001
        logger.debug("Não foi possível localizar saldo do balancete para %s: %s", conta, exc)
    return None


def _ler_valor_cache_subtotal(arquivo: str, aba: str) -> float | None:
    """Localiza a celula que contem uma formula =SUBTOTAL(...) perto do topo da
    planilha (linhas 1-10) e devolve o VALOR EM CACHE dessa mesma celula - so
    para fins de comparacao/alerta, nunca para calculo (o motor sempre soma a
    coluna Valor recalculada em Python). Devolve None se nao achar nenhuma.

    Por que isso importa: no arquivo que motivou este projeto, a formula
    =SUBTOTAL(9,M7:M120) tinha um valor em cache desatualizado (o Excel so
    recalcula ao abrir/editar - um arquivo gerado por script e nunca reaberto
    fica com cache velho para sempre). Confiar nesse cache silenciosamente
    teria produzido uma conferencia de saldo errada.
    """
    try:
        wb_formulas = load_workbook(arquivo, read_only=True, data_only=False)
        wb_cache = load_workbook(arquivo, read_only=True, data_only=True)
        ws_formulas = wb_formulas[aba]
        ws_cache = wb_cache[aba]
        for row in ws_formulas.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if isinstance(cell.value, str) and "SUBTOTAL" in cell.value.upper():
                    valor_cache = ws_cache[cell.coordinate].value
                    if isinstance(valor_cache, (int, float)):
                        logger.info(
                            "Formula de conferencia encontrada em %s: %s (cache = %s)",
                            cell.coordinate, cell.value, valor_cache,
                        )
                        return float(valor_cache)
    except Exception as exc:  # noqa: BLE001 - so um alerta auxiliar, nunca deve derrubar a leitura
        logger.debug("Nao foi possivel ler valor de cache de SUBTOTAL: %s", exc)
    return None
